"""
Product Data Scraper
====================
Scrapes product Name, Description, and Specifications from the source
e-commerce platform and writes results to the configured Excel columns.

Part of the e-commerce product data migration pipeline.
See README.md for full setup and configuration instructions.
"""

import re
import time
import logging
import random
from pathlib import Path

from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from config import get_settings
# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
app_settings = get_settings()

INPUT_EXCEL_FILE   = app_settings.EXCEL_FILE_PATH
INPUT_SHEET_NAME   = app_settings.SHEET_NAME
ITEM_NUMBER_COLUMN = app_settings.ITEM_NUMBER_COLUMN
HEADER_ROW         = 1

# ── Batch control ─────────────────────────────────────────────────────────────
START_ROW        = app_settings.START_ROW
MAX_ROWS_PER_RUN = app_settings.MAX_ROWS_PER_RUN
# ─────────────────────────────────────────────────────────────────────────────

COL_ITEM     = app_settings.ITEM_NUMBER_COLUMN
COL_NAME     = app_settings.COL_NAME_EN
COL_OVERVIEW = app_settings.COL_DESC_EN
COL_SPECS    = app_settings.COL_SPECS_EN

URL_TEMPLATE = "https://www.acehardware.com/departments/hardware/nails-and-staples/nails/{item}"

HEADLESS            = True
PAGE_TIMEOUT_MS     = 30_000
SELECTOR_TIMEOUT_MS = 10_000

MIN_DELAY = 0.8
MAX_DELAY = 1.8

# ---------------------------------------------------------------------------
# Lines to filter out
# ---------------------------------------------------------------------------
FILTERED_PREFIXES = ("click", "california")

def _should_filter_line(line: str) -> bool:
    return line.strip().lower().startswith(FILTERED_PREFIXES)

def sanitize(value: str) -> str:
    if not isinstance(value, str):
        return value
    return re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', value)

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Browser
# ---------------------------------------------------------------------------

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_5) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:126.0) Gecko/20100101 Firefox/126.0",
]


def build_browser_context(playwright):
    browser = playwright.chromium.launch(
        headless=HEADLESS,
        args=[
            "--disable-blink-features=AutomationControlled",
            "--no-sandbox",
            "--disable-dev-shm-usage",
            "--blink-settings=imagesEnabled=false",
        ],
    )
    context = browser.new_context(
        user_agent=random.choice(USER_AGENTS),
        viewport={"width": 1366, "height": 768},
        locale="en-US",
        timezone_id="America/Chicago",
        extra_http_headers={
            "Accept-Language": "en-US,en;q=0.9",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Encoding": "gzip, deflate, br",
            "Upgrade-Insecure-Requests": "1",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "none",
        },
    )
    context.add_init_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )

    # Not blocking resources — some pages need stylesheets/scripts to render specs
    return browser, context


# ---------------------------------------------------------------------------
# Scraping
# ---------------------------------------------------------------------------

def scrape_product(page, item_number: str) -> dict:
    url = URL_TEMPLATE.format(item=item_number)
    log.info("Fetching item %s", item_number)

    page.goto(url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT_MS)

    # Dismiss cookie consent popup if it appears
    try:
        page.wait_for_selector("button:has-text('Necessary Cookies Only')", timeout=5000)
        page.click("button:has-text('Necessary Cookies Only')")
        log.info("Cookie popup dismissed")
        page.wait_for_timeout(1000)
    except PlaywrightTimeout:
        pass  # no cookie popup, continue normally

    try:
        page.wait_for_selector("h1.mz-pagetitle.title", timeout=SELECTOR_TIMEOUT_MS)
    except PlaywrightTimeout:
        raise RuntimeError(f"Product title not found for item {item_number} (possible 404 or bot block)")

    name     = page.locator("h1.mz-pagetitle.title").inner_text().strip()
    overview = _extract_overview(page)

    # Scroll to bottom once to trigger lazy-loaded specs
    page.evaluate("window.scrollTo(0, document.body.scrollHeight)")

    # Wait for EITHER the accordion specs OR the original specs element — whichever appears first
    try:
        page.wait_for_selector(
            "#rds-accordion-body-accordian-specifications div.mz-productdetail-properties ul li, "
            "div.mz-productdetail-properties.menuToggle-menu ul li, "
            "div.mz-productdetail-properties ul li",
            timeout=8000
        )
    except PlaywrightTimeout:
        pass  # product may have no specs

    specs    = _extract_specs(page)

    return {"name": name, "overview": overview, "specs": specs}


def _extract_overview(page) -> str:
    data = page.evaluate("""
        () => {
            const el = document.getElementById('data-mz-preload-product');
            if (!el) return null;
            try { return JSON.parse(el.textContent); }
            catch(e) { return null; }
        }
    """)

    if not data:
        log.warning("Preloaded product JSON not found")
        return ""

    parts = []

    full_desc = data.get("content", {}).get("productFullDescription", "").strip()
    if full_desc:
        clean_lines = [
            line for line in full_desc.splitlines()
            if not _should_filter_line(line)
        ]
        clean_desc = "\n".join(clean_lines).strip()
        if clean_desc:
            parts.append(clean_desc)

    bullets = []
    for feature in data.get("sortedFeatureList", []):
        values = feature.get("values", [])
        if values:
            text = values[0].get("stringValue", "").strip()
            if text and not _should_filter_line(text):
                bullets.append(f"• {text}")

    if bullets:
        parts.append("\n".join(bullets))

    return "\n\n".join(parts)


def _extract_specs(page) -> str:
    """
    Try selectors in priority order:

    1. Accordion specs section — confirmed location in air filter and similar products:
       #rds-accordion-body-accordian-specifications ... ul li

    2. Original location used by nails/screws pages:
       div.mz-productdetail-properties.menuToggle-menu ul li

    3. Original fallback:
       div.mz-productdetail-properties ul li
    """

    # Priority 1 — accordion specs section (confirmed in HTML you shared)
    SELECTOR_1 = "#rds-accordion-body-accordian-specifications div.mz-productdetail-properties ul li"

    # Priority 2 — original selector
    SELECTOR_2 = "div.mz-productdetail-properties.menuToggle-menu ul li"

    # Priority 3 — original fallback
    SELECTOR_3 = "div.mz-productdetail-properties ul li"

    for selector in [SELECTOR_1, SELECTOR_2, SELECTOR_3]:
        lines = _collect_spec_lines(page, selector)
        if lines:
            log.debug("Specs found via: %s", selector)
            return "\n".join(lines)

    # Debug dump
    log.warning("No specs found — running debug dump...")
    for selector in [SELECTOR_1, SELECTOR_2, SELECTOR_3]:
        count = page.locator(selector).count()
        log.warning("  Selector found %d elements: %s", count, selector)
    acc = page.locator("#rds-accordion-body-accordian-specifications").count()
    props = page.locator("div.mz-productdetail-properties").count()
    log.warning("  Accordion container exists: %s", acc > 0)
    log.warning("  Any mz-productdetail-properties exists: %s", props > 0)
    if acc > 0:
        html = page.locator("#rds-accordion-body-accordian-specifications").inner_html()
        log.warning("  Accordion HTML (500 chars): %s", html[:500])
    if props > 0:
        html2 = page.locator("div.mz-productdetail-properties").first.inner_html()
        log.warning("  Properties HTML (500 chars): %s", html2[:500])
    page.screenshot(path="debug_specs.png")
    log.warning("  Screenshot saved: debug_specs.png")
    return ""


def _collect_spec_lines(page, css: str) -> list[str]:
    lines = []
    for li in page.locator(css).all():
        text = li.inner_text().strip()
        if not text:
            continue
        text = " ".join(text.split())
        if "safety data sheets" in text.lower():
            continue
        if _should_filter_line(text):
            continue
        lines.append(text)
    return lines


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def read_rows_from_excel(filepath, sheet_name, col, header_row,
                         start_data_row, max_rows):
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    col_idx = ws[f"{col}1"].column
    first_excel_row = header_row + start_data_row

    results = []
    row_num = first_excel_row
    collected = 0

    for row in ws.iter_rows(min_row=first_excel_row, min_col=col_idx,
                             max_col=col_idx, values_only=True):
        val = row[0]
        if val is not None and str(val).strip():
            results.append((row_num, str(val).strip()))
            collected += 1
            if max_rows is not None and collected >= max_rows:
                break
        row_num += 1

    wb.close()
    return results


def write_result(ws, excel_row: int, data: dict | None, error: str | None):
    wrap = Alignment(wrap_text=True, vertical="top")

    if error:
        cell = ws[f"{COL_NAME}{excel_row}"]
        cell.value = sanitize(f"ERROR: {error}")
        cell.font  = Font(color="FF0000")
    else:
        name_cell     = ws[f"{COL_NAME}{excel_row}"]
        overview_cell = ws[f"{COL_OVERVIEW}{excel_row}"]
        specs_cell    = ws[f"{COL_SPECS}{excel_row}"]

        name_cell.value     = sanitize(data["name"])
        overview_cell.value = sanitize(data["overview"])
        specs_cell.value    = sanitize(data["specs"])

        for cell in (name_cell, overview_cell, specs_cell):
            cell.alignment = wrap


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    log.info("Reading item numbers from %s  (start_row=%d, max=%s)",
             INPUT_EXCEL_FILE, START_ROW, MAX_ROWS_PER_RUN)

    rows = read_rows_from_excel(
        INPUT_EXCEL_FILE, INPUT_SHEET_NAME, ITEM_NUMBER_COLUMN,
        HEADER_ROW, START_ROW, MAX_ROWS_PER_RUN
    )

    if not rows:
        log.warning("No item numbers found. Exiting.")
        return

    log.info("Will process %d rows (Excel rows %d – %d)",
             len(rows), rows[0][0], rows[-1][0])

    wb = load_workbook(INPUT_EXCEL_FILE)
    ws = wb[INPUT_SHEET_NAME] if INPUT_SHEET_NAME in wb.sheetnames else wb.active

    with sync_playwright() as p:
        browser, context = build_browser_context(p)
        page = context.new_page()

        for excel_row, item_number in rows:
            try:
                data = scrape_product(page, item_number)
                write_result(ws, excel_row, data=data, error=None)
                log.info("✓ Row %-4d  Item %-12s  Name: %.45s",
                         excel_row, item_number, data["name"])
            except Exception as exc:
                log.error("✗ Row %-4d  Item %s failed: %s", excel_row, item_number, exc)
                write_result(ws, excel_row, data=None, error=str(exc))

            time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))

        browser.close()

    wb.save(INPUT_EXCEL_FILE)
    log.info("Saved results to %s", INPUT_EXCEL_FILE)


if __name__ == "__main__":
    main()