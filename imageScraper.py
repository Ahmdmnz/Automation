"""
Product Image Downloader
========================
Downloads product images using a two-source fallback strategy:
  1. Try the primary source using the item number
  2. If not found → fall back to a barcode lookup service
  3. If not found on either → log to the not-found report at the end

Naming convention:
    Single image  → {item_number}.jpg
    Multiple      → {item_number}(1).jpg, {item_number}(2).jpg, ...

Part of the e-commerce product data migration pipeline.
See README.md for full setup and configuration instructions.
"""

import time
import logging
import random
import requests
from pathlib import Path

from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
from openpyxl import load_workbook
from config import get_settings
# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
app_settings = get_settings()

INPUT_EXCEL_FILE   = app_settings.EXCEL_FILE_PATH
INPUT_SHEET_NAME   = app_settings.SHEET_NAME
ITEM_NUMBER_COLUMN = app_settings.ITEM_NUMBER_COLUMN   # Ace item number
BARCODE_COLUMN     = app_settings.BARCODE_COLUMN   # Barcode for fallback
HEADER_ROW         = 1

IMAGES_FOLDER = app_settings.IMAGES_FOLDER

# ── Batch control ─────────────────────────────────────────────────────────────
START_ROW        = app_settings.START_ROW    # 1 = first data row after header (Excel row 2)
MAX_ROWS_PER_RUN = app_settings.MAX_ROWS_PER_RUN   # None = all rows
# ─────────────────────────────────────────────────────────────────────────────

ACE_URL_TEMPLATE     = "https://www.acehardware.com/departments/hardware/nails-and-staples/nails/{item}"
BARCODE_URL_TEMPLATE = "https://www.barcodelookup.com/{barcode}"

HEADLESS            = True
PAGE_TIMEOUT_MS     = 30_000
SELECTOR_TIMEOUT_MS = 10_000

MIN_DELAY = 2.0
MAX_DELAY = 4.5

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Browser
# ---------------------------------------------------------------------------

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_5) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:126.0) Gecko/20100101 Firefox/126.0",
]


def build_browser_context(playwright):
    browser = playwright.chromium.launch(
        headless=HEADLESS,
        args=[
            "--disable-blink-features=AutomationControlled",
            "--no-sandbox",
            "--disable-dev-shm-usage",
        ],
    )
    context = browser.new_context(
        user_agent=random.choice(USER_AGENTS),
        viewport={"width": 1366, "height": 768},
        locale="en-US",
        timezone_id="America/Chicago",
        extra_http_headers={
            "Accept-Language": "en-US,en;q=0.9",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Encoding": "gzip, deflate, br",
            "Upgrade-Insecure-Requests": "1",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "none",
        },
    )
    context.add_init_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    return browser, context


# ---------------------------------------------------------------------------
# Ace Hardware scraping
# ---------------------------------------------------------------------------

def scrape_ace(page, item_number: str) -> list[str]:
    """
    Try to get images from acehardware.com.
    Returns list of image URLs, or empty list if not found.
    """
    url = ACE_URL_TEMPLATE.format(item=item_number)
    log.info("  [Ace] Fetching %s", url)

    page.goto(url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT_MS)

    try:
        page.wait_for_selector("h1.mz-pagetitle.title", timeout=SELECTOR_TIMEOUT_MS)
    except PlaywrightTimeout:
        log.info("  [Ace] Not found for item %s", item_number)
        return []

    data = page.evaluate("""
        () => {
            const el = document.getElementById('data-mz-preload-product');
            if (!el) return null;
            try { return JSON.parse(el.textContent); }
            catch(e) { return null; }
        }
    """)

    if not data:
        log.warning("  [Ace] Preloaded JSON not found for item %s", item_number)
        return []

    urls = []
    for img in data.get("content", {}).get("productImages", []):
        img_url = img.get("imageUrl", "")
        if img_url:
            urls.append("https:" + img_url if img_url.startswith("//") else img_url)

    log.info("  [Ace] Found %d image(s) for item %s", len(urls), item_number)
    return urls


# ---------------------------------------------------------------------------
# Barcode Lookup scraping
# ---------------------------------------------------------------------------

def scrape_barcodelookup(page, barcode: str) -> list[str]:
    """
    Try to get images from barcodelookup.com.
    Returns list of image URLs, or empty list if not found.
    """
    url = BARCODE_URL_TEMPLATE.format(barcode=barcode)
    log.info("  [Barcode] Fetching %s", url)

    page.goto(url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT_MS)

    try:
        page.wait_for_selector("#largeProductImage img", timeout=SELECTOR_TIMEOUT_MS)
    except PlaywrightTimeout:
        log.info("  [Barcode] Not found for barcode %s", barcode)
        return []

    urls = []

    # Main large image
    main_img = page.locator("#largeProductImage img").first
    src = main_img.get_attribute("src") or ""
    if src.startswith("http"):
        urls.append(src)

    # Additional thumbnails
    for thumb in page.locator("#productImageThumbs .thumb-box img").all():
        src = thumb.get_attribute("src") or ""
        if src.startswith("http") and src not in urls:
            urls.append(src)

    log.info("  [Barcode] Found %d image(s) for barcode %s", len(urls), barcode)
    return urls


# ---------------------------------------------------------------------------
# Download
# ---------------------------------------------------------------------------

def download_images(urls: list[str], name: str, folder: str):
    Path(folder).mkdir(parents=True, exist_ok=True)
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}

    if len(urls) == 1:
        file_path = Path(folder) / f"{name}.jpg"
        r = requests.get(urls[0], headers=headers, timeout=15)
        with open(file_path, "wb") as f:
            f.write(r.content)
    else:
        for index, url in enumerate(urls, start=1):
            file_path = Path(folder) / f"{name}({index}).jpg"
            r = requests.get(url, headers=headers, timeout=15)
            with open(file_path, "wb") as f:
                f.write(r.content)


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def read_rows_from_excel(filepath, sheet_name, item_col, barcode_col,
                         header_row, start_data_row, max_rows):
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    item_col_idx   = ws[f"{item_col}1"].column
    barcode_col_idx = ws[f"{barcode_col}1"].column
    first_excel_row = header_row + start_data_row

    results = []
    row_num = first_excel_row
    collected = 0

    for row in ws.iter_rows(min_row=first_excel_row, values_only=True):
        item_val    = row[item_col_idx - 1]
        barcode_val = row[barcode_col_idx - 1]

        item_str    = str(item_val).strip() if item_val is not None else ""
        barcode_str = str(barcode_val).strip() if barcode_val is not None else ""

        if item_str or barcode_str:
            results.append((row_num, item_str, barcode_str))
            collected += 1
            if max_rows is not None and collected >= max_rows:
                break
        row_num += 1

    wb.close()
    return results


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    log.info("Reading from %s (start_row=%d, max=%s)",
             INPUT_EXCEL_FILE, START_ROW, MAX_ROWS_PER_RUN)

    rows = read_rows_from_excel(
        INPUT_EXCEL_FILE, INPUT_SHEET_NAME,
        ITEM_NUMBER_COLUMN, BARCODE_COLUMN,
        HEADER_ROW, START_ROW, MAX_ROWS_PER_RUN
    )

    if not rows:
        log.warning("No rows found. Exiting.")
        return

    log.info("Will process %d rows (Excel rows %d – %d)",
             len(rows), rows[0][0], rows[-1][0])

    not_found = []   # rows where neither Ace nor Barcode found images

    with sync_playwright() as p:
        browser, context = build_browser_context(p)
        page = context.new_page()

        for excel_row, item_number, barcode in rows:
            log.info("Row %d | Item: %s | Barcode: %s", excel_row, item_number, barcode)

            image_urls = []
            source = None

            # ── Step 1: Try Ace Hardware by item number ──────────────────
            if item_number:
                try:
                    image_urls = scrape_ace(page, item_number)
                    if image_urls:
                        source = "Ace"
                except Exception as e:
                    log.error("  [Ace] Error: %s", e)

            # ── Step 2: Fallback to Barcode Lookup ───────────────────────
            if not image_urls and barcode:
                try:
                    image_urls = scrape_barcodelookup(page, barcode)
                    if image_urls:
                        source = "Barcode"
                except Exception as e:
                    log.error("  [Barcode] Error: %s", e)

            # ── Step 3: Handle result ─────────────────────────────────────
            if image_urls:
                file_name = item_number if item_number else barcode
                try:
                    download_images(image_urls, file_name, IMAGES_FOLDER)
                    log.info("  ✓ Downloaded %d image(s) via %s → %s",
                             len(image_urls), source, file_name)
                except Exception as e:
                    log.error("  ✗ Download failed: %s", e)
            else:
                log.warning("  ✗ No images found on Ace or Barcode Lookup")
                not_found.append(excel_row)

            time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))

        browser.close()

    # ── Final summary ─────────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"Total processed : {len(rows)}")
    print(f"Not found       : {len(not_found)}")
    if not_found:
        print(f"Not found rows  : {not_found}")
    print("=" * 60)


if __name__ == "__main__":
    main()