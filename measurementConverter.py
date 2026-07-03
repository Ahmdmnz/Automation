"""
Measurement Converter
=====================
Converts imperial measurements to metric equivalents inline within
product text — supporting dimensions, weight, volume, area, and temperature.

Examples:
    5 in          → 5 in (12.7 cm)
    2.5 lbs       → 2.5 lbs (1.13 kg)
    3 x 5 x 2 in  → 3 x 5 x 2 in (7.62 x 12.7 x 5.08 cm)

Safe to re-run — already-converted values are never double-converted.

Part of the e-commerce product data migration pipeline.
See README.md for full setup and configuration instructions.
"""

import re
from pathlib import Path
from openpyxl import load_workbook
from config import get_settings
# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
app_settings = get_settings()

EXCEL_FILE_PATH = app_settings.EXCEL_FILE_PATH
SHEET_NAME      = app_settings.SHEET_NAME
COLS_TO_PROCESS = [
    app_settings.COL_NAME_EN,
    app_settings.COL_DESC_EN,
    app_settings.COL_SPECS_EN,
    app_settings.COL_META_TITLE_EN,
    app_settings.COL_META_DESC_EN,
    app_settings.COL_TAGS_EN,
    app_settings.COL_BRAND_EN,
]
START_ROW = app_settings.START_ROW

# ---------------------------------------------------------------------------
# Catalogue-prefix detector
# ---------------------------------------------------------------------------
_CATALOGUE_PREFIX = re.compile(
    r'(?:No\.?|#|Gauge|Ga\.?|Size|Qty|Pack|Pk\.?)\s*$',
    re.IGNORECASE,
)

# ---------------------------------------------------------------------------
# Fraction / mixed-number parser
# ---------------------------------------------------------------------------

def fraction_to_float(s: str) -> float:
    s = s.strip()
    try:
        if '-' in s and '/' in s:
            whole, frac = s.split('-', 1)
            num, den = frac.split('/')
            return float(whole) + float(num) / float(den)
        elif ' ' in s and '/' in s:
            whole, frac = s.rsplit(' ', 1)
            num, den = frac.split('/')
            return float(whole) + float(num) / float(den)
        elif '/' in s:
            num, den = s.split('/')
            return float(num) / float(den)
        else:
            return float(s)
    except Exception:
        return 0.0

# ---------------------------------------------------------------------------
# Core converter
# ---------------------------------------------------------------------------

def _fmt(v: float) -> str:
    return f"{v:.2f}".rstrip('0').rstrip('.')

def _convert(val_str: str, unit_str: str, original_text: str) -> str | None:
    u = unit_str.strip().lower()
    parts = re.split(r'\s*[xX*]\s*', val_str)
    vals  = [fraction_to_float(p) for p in parts]

    def apply(factor, label):
        metric = [v * factor for v in vals]
        return f"{original_text} ({' x '.join(_fmt(v) for v in metric)} {label})"

    # Area
    if re.match(r'sq\.?\s*(?:ft\.?|feet|foot)|square\s*(?:feet|foot|ft\.?)', u):
        return apply(0.092903, "m²")
    if re.match(r'sq\.?\s*(?:in\.?|inches?|inch)|square\s*(?:in\.?|inches?|inch)', u):
        return apply(6.4516, "cm²")

    # Length
    if re.fullmatch(r'in\.?|inch|inches|"', u):
        return apply(2.54, "cm")
    if re.fullmatch(r"ft\.?|foot|feet|'", u):
        return apply(0.3048, "m")

    # Weight
    if re.fullmatch(r'lbs?\.?|pounds?', u):
        return apply(0.453592, "kg")

    # Fluid volume (before plain oz)
    if re.fullmatch(r'fl\.?\s*oz\.?|fl\.?\s*ounces?', u):
        return apply(29.5735, "ml")

    # Dry oz
    if re.fullmatch(r'oz\.?|ounces?', u):
        return apply(28.3495, "g")

    # Volume
    if re.fullmatch(r'gals?\.?|gallons?', u):
        return apply(3.78541, "L")

    # Temperature
    if re.fullmatch(r'degrees?\s*fahrenheit|fahrenheit|°[fF]', u):
        metric = [(v - 32) * 5.0 / 9.0 for v in vals]
        return f"{original_text} ({' x '.join(_fmt(v) for v in metric)} °C)"

    return None

# ---------------------------------------------------------------------------
# Regex
#
# KEY CHANGE: removed \b before unit keywords.
# Instead, each unit token ends with a negative lookahead (?![a-zA-Z])
# so "finishing" won't match (f-i-n-i-s-h-i-n-g has letters after "in")
# but "12in", "5 in", "3in." all match because after "in" comes a digit,
# space, punctuation, or end-of-string — not a letter.
#
# The chain _CHAIN already requires a number immediately before the unit,
# so "in kitchen" never matches ("kitchen" is not a number).
# ---------------------------------------------------------------------------

_N = r'(?:-?\d+\s+\d+/\d+|-?\d+-\d+/\d+|-?\d+/\d+|-?\d+\.\d+|-?\.\d+|-?\d+)'
_CHAIN = r'(' + _N + r'(?:\s*[xX*]\s*' + _N + r')*)'

# Unit — ends with (?![a-zA-Z]) to prevent mid-word matches
_U = r'(?![a-zA-Z])'  # negative lookahead: no letter right after the unit

_UNIT = (
    r'('
    r'sq\.?\s*(?:ft\.?|feet|foot|in\.?|inches?|inch)' + _U +
    r'|square\s*(?:feet|foot|ft\.?|in\.?|inches?|inch)' + _U +
    r'|feet' + _U + r'|foot' + _U + r'|ft\.?' + _U +
    r'|inches?' + _U + r'|inch' + _U +
    r'|in\.?' + _U +           # "in" / "in." — no letter after
    r'|fl\.?\s*oz\.?' + _U + r'|fl\.?\s*ounces?' + _U +
    r'|pounds?' + _U + r'|lbs?\.?' + _U +
    r'|ounces?' + _U + r'|oz\.?' + _U +
    r'|gallons?' + _U + r'|gals?\.?' + _U +
    r'|degrees?\s*fahrenheit' + _U + r'|fahrenheit' + _U + r'|°[Ff]' +
    r'|"|\''
    r')'
)

_NO_DUP = r'(?!\s*\(\s*[\d\.\s\-x]+\s*(?:cm|m|kg|ml|g|°C|cm²|m²|L)\s*\))'

MEASUREMENT_PATTERN = re.compile(
    _CHAIN + r'\s*' + _UNIT + _NO_DUP,
    re.IGNORECASE,
)

# ---------------------------------------------------------------------------
# Text processor
# ---------------------------------------------------------------------------

def process_text(text: str) -> str:
    if not text or not isinstance(text, str):
        return text

    def replacer(match):
        original = match.group(0)
        val_str  = match.group(1).strip()
        unit_str = match.group(2).strip()

        parts = re.split(r'\s*[xX*]\s*', val_str)

        if len(parts) > 1:
            before = text[max(0, match.start() - 25): match.start()]
            if _CATALOGUE_PREFIX.search(before):
                real_val = parts[-1].strip()
                sep_pattern = re.compile(r'\s*[xX*]\s*' + re.escape(real_val))
                m2 = sep_pattern.search(original)
                if m2:
                    prefix_part = original[:m2.start()]
                    measurable  = real_val + original[m2.start() + len(m2.group(0)):]
                    converted   = _convert(real_val, unit_str, measurable)
                    if converted:
                        return prefix_part + " x " + converted
                return original

        converted = _convert(val_str, unit_str, original)
        return converted if converted is not None else original

    return MEASUREMENT_PATTERN.sub(replacer, text)

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if not Path(EXCEL_FILE_PATH).exists():
        print(f"File not found: {EXCEL_FILE_PATH}")
        return

    print("Loading workbook...")
    wb = load_workbook(EXCEL_FILE_PATH)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    processed_count = 0
    for row in range(START_ROW, ws.max_row + 1):
        modified_row = False
        for col in COLS_TO_PROCESS:
            cell_value = ws[f"{col}{row}"].value
            if cell_value and isinstance(cell_value, str):
                new_value = process_text(cell_value)
                if new_value != cell_value:
                    ws[f"{col}{row}"].value = new_value
                    modified_row = True
        if modified_row:
            processed_count += 1

    wb.save(EXCEL_FILE_PATH)
    print(f"Done! Updated {processed_count} rows.")


# ---------------------------------------------------------------------------
# Self-tests
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    tests = [
        # No-space versions — must convert
        ("16oz",                         "16oz (453.59 g)"),
        ("12in",                         "12in (30.48 cm)"),
        ("5ft",                          "5ft (1.52 m)"),
        ("2.5lbs",                       "2.5lbs (1.13 kg)"),
        ("1/2in",                        "1/2in (1.27 cm)"),
        ("28fl.oz",                      "28fl.oz (828.06 ml)"),
        # Spaced versions — must convert
        ("16 oz",                        "16 oz (453.59 g)"),
        ("12 in",                        "12 in (30.48 cm)"),
        ("5 ft",                         "5 ft (1.52 m)"),
        ("1/2 in",                       "1/2 in (1.27 cm)"),
        # Dimension chains
        ("3 x 5 x 2 in",                "3 x 5 x 2 in (7.62 x 12.7 x 5.08 cm)"),
        # Catalogue prefix — only last number converts
        ("No. 4 x 5 in",                "No. 4 x 5 in (12.7 cm)"),
        ("Gauge 8 x 3 in",              "Gauge 8 x 3 in (7.62 cm)"),
        ("# 10 x 2 in",                 "# 10 x 2 in (5.08 cm)"),
        # Preposition 'in' — must NOT convert
        ("available in 5 colors",       "available in 5 colors"),
        ("comes in 3 sizes",            "comes in 3 sizes"),
        ("used in outdoor",             "used in outdoor"),
        # Mid-word — must NOT convert
        ("finishing nails",             "finishing nails"),
        ("including hardware",          "including hardware"),
        ("contains 4 items",            "contains 4 items"),
        # Anti-duplicate
        ("5 in (12.7 cm)",              "5 in (12.7 cm)"),
    ]

    print("Running self-tests...\n")
    all_pass = True
    for inp, expected in tests:
        result = process_text(inp)
        ok     = result == expected
        if not ok:
            all_pass = False
        print(f'  {"✓" if ok else "✗"}  {inp!r}')
        if not ok:
            print(f'      EXP: {expected!r}')
            print(f'      GOT: {result!r}')

    print()
    if all_pass:
        print("All tests passed — running on Excel file...\n")
        main()
    else:
        print("Some tests failed — NOT running on Excel file.")