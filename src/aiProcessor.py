"""
AI Metadata Processor
=====================
Uses Groq AI (LLaMA 3.3 70B) to generate SEO-optimized English metadata
for each product based on its name, description, and specifications:
    - Meta Title       (max 60 characters)
    - Meta Description (max 160 characters)
    - Search Tags      (8-12 comma-separated terms)
    - Brand Name       (extracted from product data)

Automatically rotates API keys on rate limit to keep the pipeline running.
Skips already-processed rows — safe to resume after interruption.

Part of the e-commerce product data migration pipeline.
See README.md for full setup and configuration instructions.
"""

import json
import re
import time
import logging
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from groq import Groq
from config import get_settings
# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
app_settings = get_settings()
EXCEL_FILE_PATH = app_settings.EXCEL_FILE_PATH
SHEET_NAME      = app_settings.SHEET_NAME

# ── Input columns (scraped data) ──────────────────────────────────────────────
COL_NAME_EN  = app_settings.COL_NAME_EN
COL_DESC_EN  = app_settings.COL_DESC_EN
COL_SPECS_EN = app_settings.COL_SPECS_EN

# ── Output columns ─────────────────────────────────────────────────────────────
COL_META_TITLE_EN = app_settings.COL_META_TITLE_EN
COL_META_DESC_EN  = app_settings.COL_META_DESC_EN
COL_TAGS_EN       = app_settings.COL_TAGS_EN
COL_BRAND_EN      = app_settings.COL_BRAND_EN

OUTPUT_COLS = [COL_META_TITLE_EN, COL_META_DESC_EN, COL_TAGS_EN, COL_BRAND_EN]

# ── Groq ───────────────────────────────────────────────────────────────────────
AI_MODEL = app_settings.AI_MODEL

API_KEYS = app_settings.GROQ_API_KEYS

current_key_index = 0

def get_client():
    return Groq(api_key=API_KEYS[current_key_index])

# ── Batch control ──────────────────────────────────────────────────────────────
START_ROW        = app_settings.START_ROW     # row 1 is header
MAX_ROWS_PER_RUN = app_settings.MAX_ROWS_PER_RUN  # None = all rows
# ─────────────────────────────────────────────────────────────────────────────

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Sanitize
# ---------------------------------------------------------------------------

def sanitize(value: str) -> str:
    if not isinstance(value, str):
        return value
    return re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', value)

# ---------------------------------------------------------------------------
# Prompt — English only
# ---------------------------------------------------------------------------

SYSTEM_PROMPT = """You are an expert e-commerce catalog specialist for a hardware store.
Given a product's English Name, Description, and Specifications, generate the following English fields.

Rules:
1. Meta Title: concise, SEO-friendly, under 60 characters.
2. Meta Description: engaging, SEO-friendly, under 160 characters.
3. Tags: comma-separated relevant search terms (8–12 tags).
4. Brand: extract the manufacturer/brand name only (e.g. "HILLMAN", "3M", "DAP"). If not identifiable, return "".

You MUST respond with a single valid JSON object and nothing else — no explanation, no markdown, no code fences:
{
  "meta_title_en": "SEO Meta Title (max 60 chars)",
  "meta_description_en": "SEO Meta Description (max 160 chars)",
  "tags_en": "tag1, tag2, tag3",
  "brand_name_en": "BRAND"
}"""


# ---------------------------------------------------------------------------
# AI call with key rotation
# ---------------------------------------------------------------------------

def process_row_with_ai(name_en: str, desc_en: str, specs_en: str) -> dict:
    global current_key_index

    user_message = json.dumps(
        {
            "name_en":           name_en  or "",
            "description_en":    desc_en  or "",
            "specifications_en": specs_en or "",
        },
        ensure_ascii=False,
    )

    for attempt in range(len(API_KEYS)):
        try:
            client = get_client()
            response = client.chat.completions.create(
                model=AI_MODEL,
                messages=[
                    {"role": "system", "content": SYSTEM_PROMPT},
                    {"role": "user",   "content": user_message},
                ],
                temperature=0.2,
                max_tokens=512,
                response_format={"type": "json_object"},
            )
            raw = response.choices[0].message.content.strip()
            # Strip accidental markdown fences
            if raw.startswith("```"):
                raw = raw.split("\n", 1)[-1]
            if raw.endswith("```"):
                raw = raw.rsplit("```", 1)[0]
            return json.loads(raw.strip())

        except Exception as e:
            err = str(e)
            if "429" in err or "rate" in err.lower() or "limit" in err.lower():
                log.warning("Rate limit on key %d — rotating to next key.", current_key_index)
                current_key_index = (current_key_index + 1) % len(API_KEYS)
                time.sleep(2)
            else:
                raise

    raise RuntimeError("All API keys exhausted / rate limited.")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if not Path(EXCEL_FILE_PATH).exists():
        log.error("File not found: %s", EXCEL_FILE_PATH)
        return

    log.info("Loading workbook: %s", EXCEL_FILE_PATH)
    wb = load_workbook(EXCEL_FILE_PATH)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    total_rows = ws.max_row
    last_row   = min(START_ROW + MAX_ROWS_PER_RUN - 1, total_rows) if MAX_ROWS_PER_RUN else total_rows

    log.info("Processing rows %d → %d", START_ROW, last_row)

    wrap = Alignment(wrap_text=True, vertical="top")

    try:
        for row in range(START_ROW, last_row + 1):
            name_en  = ws[f"{COL_NAME_EN}{row}"].value
            desc_en  = ws[f"{COL_DESC_EN}{row}"].value
            specs_en = ws[f"{COL_SPECS_EN}{row}"].value

            # Skip empty or errored rows
            if not name_en or "ERROR:" in str(name_en):
                continue

            # Skip already processed rows
            if ws[f"{COL_META_TITLE_EN}{row}"].value:
                log.info("Row %d — already processed, skipping.", row)
                continue

            log.info("Processing row %d/%d | %s", row, last_row, str(name_en)[:55])

            try:
                result = process_row_with_ai(
                    str(name_en),
                    str(desc_en)  if desc_en  else "",
                    str(specs_en) if specs_en else "",
                )

                ws[f"{COL_META_TITLE_EN}{row}"].value = sanitize(result.get("meta_title_en", ""))
                ws[f"{COL_META_DESC_EN}{row}"].value  = sanitize(result.get("meta_description_en", ""))
                ws[f"{COL_TAGS_EN}{row}"].value       = sanitize(result.get("tags_en", ""))
                ws[f"{COL_BRAND_EN}{row}"].value      = sanitize(result.get("brand_name_en", ""))

                for col in OUTPUT_COLS:
                    ws[f"{col}{row}"].alignment = wrap

                wb.save(EXCEL_FILE_PATH)
                log.info("✓ Row %d saved.", row)

                time.sleep(0.5)

            except json.JSONDecodeError as e:
                log.error("✗ Row %d — invalid JSON: %s", row, e)
            except Exception as e:
                log.error("✗ Row %d failed: %s", row, e)

    except KeyboardInterrupt:
        log.warning("Interrupted — saving progress...")
    except PermissionError:
        log.error("Cannot save — close the Excel file first, then re-run.")
    finally:
        try:
            wb.save(EXCEL_FILE_PATH)
            log.info("Done. Saved: %s", EXCEL_FILE_PATH)
        except PermissionError:
            log.error("Final save failed — Excel file is still open!")


if __name__ == "__main__":
    main()